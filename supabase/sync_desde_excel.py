"""
Sincroniza el catalogo inicial desde el Excel de Drive hacia Supabase.

FUENTE DE DATOS (solo al inicio):
  Excel: G:\\Mi unidad\\catalogo_inicial.xlsx  -> hoja "Productos" -> TABLA "TablaProductos"
  Columnas de la tabla: SKU | PRODUCTO | VENTA $ | CATEGORIA | UND VENTA
  - VENTA $ es USD directo (se guarda en precio_usd sin conversion).
  - La columna "Imagenes" de la hoja es CHECKLIST PERSONAL del usuario y SE IGNORA.

IMAGENES:
  Carpeta Drive: G:\\Mi unidad\\puntoVenta2Tabla\\imagenes\\{SKU}.webp
  Se suben a Supabase Storage bucket "productos" en {empresa_id}/{producto_id}.webp
  y se enlazan via producto.imagen_url.

  MODO POR DEFECTO (seguro): NUNCA se pisa una imagen_url existente en Supabase.
  El .webp de Drive solo se usa si el producto NO tiene imagen_url.
  Flag --forzar-drive: sube y sobreescribe imagen_url con el .webp de Drive aunque existiera.

REQUISITOS:
  pip install supabase openpyxl
  Variables de entorno (NO hardcodeadas):
    SUPABASE_URL            ej. https://pvopcajqersioqlmccwg.supabase.co
    SUPABASE_SERVICE_ROLE   secret key (bypasea RLS y storage policies)

USO:
  $env:SUPABASE_URL="..."
  $env:SUPABASE_SERVICE_ROLE="..."
  py supabase/sync_desde_excel.py [--forzar-drive]

ADVERTENCIA: Este script es para cargar/configurar el catalogo inicial. Una vez que
la app esta en uso, el catalogo se gestiona desde la app, no desde este script.
"""
import os
import sys
import glob

try:
    from supabase import create_client, Client
except ImportError:
    sys.exit("Falta supabase: pip install supabase")

try:
    import openpyxl
    from openpyxl.utils import range_boundaries
except ImportError:
    sys.exit("Falta openpyxl: pip install openpyxl")

# ---- Config ----
EXCEL = r"G:\Mi unidad\puntoVenta2Tabla\catalogo_inicial.xlsx"
HOJA = "Productos"
TABLA = "TablaProductos"
IMG_DIR = r"G:\Mi unidad\puntoVenta2Tabla\imagenes"
EMPRESA_NOMBRE = "FerrehogarMart"
BUCKET = "productos"

URL = os.environ.get("SUPABASE_URL")
KEY = os.environ.get("SUPABASE_SERVICE_ROLE")
if not URL or not KEY:
    sys.exit("Defini SUPABASE_URL y SUPABASE_SERVICE_ROLE como variables de entorno.")

FORZAR_DRIVE = "--forzar-drive" in sys.argv

supabase: Client = create_client(URL, KEY)


def _safe(res):
    if res is None:
        raise RuntimeError("Respuesta None de Supabase (¿URL/key malas o red?)")
    if getattr(res, "error", None):
        raise RuntimeError(f"Error Supabase: {res.error}")
    return res


def get_empresa():
    res = _safe(supabase.table("empresa").select("id,nombre").eq("nombre", EMPRESA_NOMBRE).limit(1).execute())
    if not res.data:
        sys.exit(f"Empresa '{EMPRESA_NOMBRE}' no encontrada.")
    print(f"[empresa] {res.data[0]['id']}")
    return res.data[0]["id"]


def get_or_create_categoria(empresa_id, nombre):
    nombre = (nombre or "SIN CATEGORIA").strip().upper()
    res = _safe(supabase.table("categoria").select("id").eq("empresa_id", str(empresa_id)).eq("nombre", nombre).limit(1).execute())
    if res.data:
        return res.data[0]["id"]
    res = _safe(supabase.table("categoria").insert({"empresa_id": str(empresa_id), "nombre": nombre}).select("id").execute())
    print(f"[categoria] creada: {nombre}")
    return res.data[0]["id"]


def leer_tabla():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb[HOJA]
    if TABLA not in ws.tables:
        sys.exit(f"La tabla '{TABLA}' no existe en la hoja '{HOJA}'.")
    ref = ws.tables[TABLA].ref
    minc, minr, maxc, maxr = range_boundaries(ref)
    filas = []
    for rr in range(minr + 1, maxr + 1):
        sku = ws.cell(row=rr, column=minc).value
        if sku is None or str(sku).strip() == "":
            continue
        fila = {
            "sku": str(sku).strip(),
            "nombre": str(ws.cell(row=rr, column=minc + 1).value or "").strip() or str(sku).strip(),
            "precio_usd": float(ws.cell(row=rr, column=minc + 2).value) if ws.cell(row=rr, column=minc + 2).value else 0.0,
            "categoria": str(ws.cell(row=rr, column=minc + 3).value or "SIN CATEGORIA").strip().upper(),
            "unidad": str(ws.cell(row=rr, column=minc + 4).value or "UND").strip().upper(),
        }
        filas.append(fila)

    # ---- Validacion anti-duplicados ----
    # El SKU es la identidad del producto y debe ser unico. Si hay repetidos,
    # abortamos SIN tocar Supabase para no encubrir el error ni romper el upsert.
    from collections import Counter
    conteo = Counter(f["sku"] for f in filas)
    duplicados = {k: n for k, n in conteo.items() if n > 1}
    if duplicados:
        print("[ERROR] SKUs duplicados detectados en la tabla:")
        for k, n in sorted(duplicados.items()):
            print(f"  {k}  x{n}")
        sys.exit(
            "Hay codigos duplicados en el Excel. Corregilos (borra la fila repetida "
            "o renombra uno) y volve a correr. No se modifico Supabase."
        )

    return filas


def main():
    empresa_id = get_empresa()
    filas = leer_tabla()
    print(f"[excel] {len(filas)} productos en {TABLA}")

    insertados = 0
    actualizados = 0
    sin_cambios = 0
    con_imagen_subida = 0
    ya_tenian = 0
    sin_imagen_local = 0
    skus_sin_match = 0

    cat_cache = {}

    def _precio_diffiere(a, b):
        # a = numeric de Postgres (puede venir como float o str), b = float del Excel
        try:
            av = float(a) if a is not None else 0.0
        except (TypeError, ValueError):
            av = 0.0
        return abs(av - b) > 0.001

    for f in filas:
        if f["categoria"] not in cat_cache:
            cat_cache[f["categoria"]] = get_or_create_categoria(empresa_id, f["categoria"])

        cat_id = str(cat_cache[f["categoria"]])
        exist = _safe(supabase.table("producto").select(
            "id,imagen_url,nombre,precio_usd,categoria_id,unidad"
        ).eq("empresa_id", str(empresa_id)).eq("sku", f["sku"]).limit(1).execute())

        if exist.data:
            p = exist.data[0]
            pid = p["id"]
            # Comparar campo a campo: solo UPDATE si algo cambió de verdad.
            cambios = {}
            if (p.get("nombre") or "") != f["nombre"]:
                cambios["nombre"] = f["nombre"]
            if _precio_diffiere(p.get("precio_usd"), f["precio_usd"]):
                cambios["precio_usd"] = f["precio_usd"]
            if (p.get("categoria_id") or "") != cat_id:
                cambios["categoria_id"] = cat_id
            if (p.get("unidad") or "").upper() != f["unidad"]:
                cambios["unidad"] = f["unidad"]

            if cambios:
                # No pisar stock_actual/stock_minimo: se mantienen en la BD.
                _safe(supabase.table("producto").update(cambios).eq("id", pid).execute())
                actualizados += 1
                print(f"[update] {f['sku']}: {list(cambios.keys())}")
            else:
                sin_cambios += 1
        else:
            res = _safe(supabase.table("producto").insert({
                "empresa_id": str(empresa_id),
                "sku": f["sku"],
                "nombre": f["nombre"],
                "categoria_id": cat_id,
                "unidad": f["unidad"],
                "precio_usd": f["precio_usd"],
                "stock_actual": 0,
                "stock_minimo": 0,
                "activo": True,
            }).select("id,imagen_url").execute())
            pid = res.data[0]["id"]
            insertados += 1

        # ---- Imagen ----
        imagen_url_actual = exist.data[0]["imagen_url"] if exist.data else None
        archivo = os.path.join(IMG_DIR, f"{f['sku']}.webp")

        if imagen_url_actual and not FORZAR_DRIVE:
            ya_tenian += 1
            continue

        if not os.path.exists(archivo):
            sin_imagen_local += 1
            continue

        with open(archivo, "rb") as fh:
            data = fh.read()
        # Convencion actual: imagenes en la RAIZ del bucket (productos/{sku}.webp),
        # igual que el seed original y que las imagen_url existentes.
        dest = f"{f['sku']}.webp"
        _safe(supabase.storage.from_(BUCKET).upload(dest, data, {"content-type": "image/webp", "upsert": "true"}))
        public_url = f"{URL}/storage/v1/object/public/{BUCKET}/{dest}"
        _safe(supabase.table("producto").update({"imagen_url": public_url}).eq("id", pid).execute())
        con_imagen_subida += 1
        print(f"[img] {f['sku']} -> {dest}" + (" (forzado)" if FORZAR_DRIVE else ""))

    print(f"[fin] insertados={insertados}, actualizados={actualizados}, sin_cambios={sin_cambios}, "
          f"imagen_subida={con_imagen_subida}, ya_tenian_imagen={ya_tenian}, "
          f"sin_imagen_local={sin_imagen_local}, skus_sin_match={skus_sin_match}")


if __name__ == "__main__":
    main()
